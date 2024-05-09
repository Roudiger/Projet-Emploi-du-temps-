import random
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

# Données initiales
matieres = ['Mathématiques', 'Physique', 'Chimie', 'Informatique', 'Biologie', 'Histoire', 'Géographie']
salles = ['A1', 'A2', 'A3', 'B1', 'B2', 'B3', 'C1', 'C2', 'C3']
profs = ['Dupont', 'Dupond', 'Martin', 'Durand', 'Lefebvre', 'Germain', 'Rousseau']
jours = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']
heures = [(i, i+1) for i in range(8, 18)]

# Création du workbook et de la feuille
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Emploi du temps'

# Ajout des en-têtes
ws['A1'] = 'Horaires'
ws.append([''] + jours)  # Ajouter une liste avec une chaîne vide et les jours

# Style des en-têtes
header_font = Font(bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = header_alignment

# Génération de l'emploi du temps
emploi_du_temps = {jour: [] for jour in jours}

for jour in jours:
    for heure in heures:
        matiere = random.choice(matieres)
        salle = random.choice(salles)
        prof = random.choice(profs)
        cours = f"{matiere} - {salle} - {prof}"
        emploi_du_temps[jour].append(cours)

# Ajout des données dans la feuille
for i, heure in enumerate(heures):
    row = [f"{heure[0]}h-{heure[1]}h"]
    for jour in jours:
        row.append(emploi_du_temps[jour][i])
    ws.append(row)

# Largeur des colonnes
ws.column_dimensions['A'].width = 12
for col in range(2, len(jours) + 2):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

# Enregistrement du fichier Excel
now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
filename = f'emploi_du_temps_{now}.xlsx'
wb.save(filename)
print(f'Le fichier {filename} a été créé avec succès.')
